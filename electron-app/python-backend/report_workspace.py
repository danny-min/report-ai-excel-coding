"""
报表工作区 (Report Workspace)
核心概念：报表是一个连续的过程，不是一次性操作
"""
import os
import uuid
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import pandas as pd
import numpy as np
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


class ReportWorkspace:
    """
    报表工作区
    - 维护当前正在编辑的报表
    - 跟踪所有 Sheet 和操作历史
    - 支持连续操作
    """
    
    def __init__(self, report_id: str = None, name: str = "未命名报表"):
        self.report_id = report_id or str(uuid.uuid4())[:8]
        self.name = name
        self.created_at = datetime.now()
        
        # 工作簿路径
        self.workbook_dir = Path("./temp/workspaces")
        self.workbook_dir.mkdir(parents=True, exist_ok=True)
        self.workbook_path = self.workbook_dir / f"{self.report_id}.xlsx"
        
        # Sheet 管理
        self.sheets: Dict[str, Dict[str, Any]] = {}
        self.active_sheet: Optional[str] = None
        
        # 操作历史
        self.history: List[Dict[str, Any]] = []
        
        # 创建空工作簿
        if not self.workbook_path.exists():
            wb = Workbook()
            wb.save(str(self.workbook_path))
    
    def add_sheet(self, name: str, df: pd.DataFrame, origin: str = "import", 
                  source_tables: List[str] = None, set_active: bool = True) -> str:
        """
        添加新 Sheet 到工作簿
        
        Args:
            name: Sheet 名称
            df: 数据 DataFrame
            origin: 来源类型 (import/merge/filter/groupby/transform)
            source_tables: 来源表名列表
            set_active: 是否设为当前活动 Sheet
        
        Returns:
            实际使用的 Sheet 名称（可能因重名而调整）
        """
        wb = load_workbook(str(self.workbook_path))
        
        # 处理重名
        final_name = name
        counter = 1
        while final_name in wb.sheetnames:
            final_name = f"{name}_{counter}"
            counter += 1
        
        # 如果是第一个 Sheet，删除默认的空 Sheet
        if len(wb.sheetnames) == 1 and wb.sheetnames[0] == "Sheet" and len(list(wb.active.iter_rows())) <= 1:
            del wb["Sheet"]
        
        # 创建新 Sheet
        ws = wb.create_sheet(title=final_name)
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(str(self.workbook_path))
        
        # 记录 Sheet 元信息
        self.sheets[final_name] = {
            "name": final_name,
            "origin": origin,
            "source_tables": source_tables or [],
            "created_at": datetime.now().isoformat(),
            "columns": df.columns.tolist(),
            "row_count": len(df),
        }
        
        # 设为活动 Sheet
        if set_active:
            self.active_sheet = final_name
        
        # 记录历史
        self.history.append({
            "action": "add_sheet",
            "sheet_name": final_name,
            "origin": origin,
            "time": datetime.now().isoformat(),
        })
        
        return final_name
    
    def get_sheet_df(self, sheet_name: str = None) -> pd.DataFrame:
        """
        获取指定 Sheet 的 DataFrame
        如果不指定，返回当前活动 Sheet
        """
        target_sheet = sheet_name or self.active_sheet
        
        if not target_sheet:
            raise ValueError("没有活动的 Sheet")
        
        if target_sheet not in self.sheets:
            raise ValueError(f"Sheet '{target_sheet}' 不存在")
        
        return pd.read_excel(str(self.workbook_path), sheet_name=target_sheet)
    
    def update_sheet(self, df: pd.DataFrame, sheet_name: str = None):
        """
        更新指定 Sheet 的数据
        """
        target_sheet = sheet_name or self.active_sheet
        
        if not target_sheet:
            raise ValueError("没有活动的 Sheet")
        
        wb = load_workbook(str(self.workbook_path))
        
        if target_sheet not in wb.sheetnames:
            raise ValueError(f"Sheet '{target_sheet}' 不存在")
        
        # 删除旧 Sheet，创建新的
        ws_idx = wb.sheetnames.index(target_sheet)
        del wb[target_sheet]
        ws = wb.create_sheet(title=target_sheet, index=ws_idx)
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        wb.save(str(self.workbook_path))
        
        # 更新元信息
        self.sheets[target_sheet]["columns"] = df.columns.tolist()
        self.sheets[target_sheet]["row_count"] = len(df)
        
        self.history.append({
            "action": "update_sheet",
            "sheet_name": target_sheet,
            "time": datetime.now().isoformat(),
        })
    
    def switch_sheet(self, sheet_name: str):
        """切换活动 Sheet"""
        if sheet_name not in self.sheets:
            raise ValueError(f"Sheet '{sheet_name}' 不存在")
        
        self.active_sheet = sheet_name
        
        self.history.append({
            "action": "switch_sheet",
            "sheet_name": sheet_name,
            "time": datetime.now().isoformat(),
        })
    
    def export(self, file_name: str = None) -> str:
        """
        导出当前报表
        
        Returns:
            导出文件的下载路径
        """
        export_dir = Path("./temp/exports")
        export_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        final_name = file_name or f"{self.name}_{timestamp}"
        if not final_name.endswith('.xlsx'):
            final_name += '.xlsx'
        
        export_path = export_dir / final_name
        shutil.copy(str(self.workbook_path), str(export_path))
        
        self.history.append({
            "action": "export",
            "file_name": final_name,
            "time": datetime.now().isoformat(),
        })
        
        return f"/api/download/{final_name}"
    
    def get_status(self) -> Dict[str, Any]:
        """获取工作区状态"""
        return {
            "report_id": self.report_id,
            "name": self.name,
            "active_sheet": self.active_sheet,
            "sheets": list(self.sheets.keys()),
            "sheets_info": self.sheets,
            "history_count": len(self.history),
        }
    
    def to_dict(self) -> Dict[str, Any]:
        """序列化为字典"""
        return {
            "report_id": self.report_id,
            "name": self.name,
            "created_at": self.created_at.isoformat(),
            "workbook_path": str(self.workbook_path),
            "active_sheet": self.active_sheet,
            "sheets": self.sheets,
            "history": self.history[-20:],  # 只保留最近20条
        }


# 全局工作区管理器
class WorkspaceManager:
    """工作区管理器"""
    
    def __init__(self):
        self.workspaces: Dict[str, ReportWorkspace] = {}
        self.current_workspace_id: Optional[str] = None
    
    def create_workspace(self, name: str = "未命名报表") -> ReportWorkspace:
        """创建新工作区"""
        ws = ReportWorkspace(name=name)
        self.workspaces[ws.report_id] = ws
        self.current_workspace_id = ws.report_id
        return ws
    
    def get_current(self) -> Optional[ReportWorkspace]:
        """获取当前工作区"""
        if self.current_workspace_id and self.current_workspace_id in self.workspaces:
            return self.workspaces[self.current_workspace_id]
        return None
    
    def get_or_create_current(self, name: str = "未命名报表") -> ReportWorkspace:
        """获取当前工作区，如果不存在则创建"""
        current = self.get_current()
        if not current:
            current = self.create_workspace(name)
        return current
    
    def switch_workspace(self, report_id: str):
        """切换工作区"""
        if report_id not in self.workspaces:
            raise ValueError(f"工作区 '{report_id}' 不存在")
        self.current_workspace_id = report_id
    
    def list_workspaces(self) -> List[Dict[str, Any]]:
        """列出所有工作区"""
        return [ws.get_status() for ws in self.workspaces.values()]


# 全局实例
workspace_manager = WorkspaceManager()





















