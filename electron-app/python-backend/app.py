"""
AI 报表生成器 - Python 后端
基于 FastAPI 的 RESTful API 服务
"""

import os
import sys
import json
import traceback
from typing import Dict, Any, List, Optional
from pathlib import Path

# 设置编码（Windows 兼容）
if sys.platform == 'win32':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse
from pydantic import BaseModel
import pandas as pd
import numpy as np
import requests
import math

# 导入自定义模块
from operation_validator import OperationValidator, format_clarification_message, get_clarification_response
from report_workspace import ReportWorkspace, WorkspaceManager, workspace_manager


def safe_json_value(val):
    """将值转换为 JSON 安全的格式"""
    if val is None:
        return None
    if isinstance(val, float):
        if math.isnan(val) or math.isinf(val):
            return None
        return val
    if isinstance(val, (np.floating, np.integer)):
        if np.isnan(val) or np.isinf(val):
            return None
        return float(val) if isinstance(val, np.floating) else int(val)
    if isinstance(val, np.ndarray):
        return val.tolist()
    if pd.isna(val):
        return None
    return val


def safe_dataframe_to_dict(df: pd.DataFrame) -> list:
    """将 DataFrame 转换为 JSON 安全的字典列表"""
    # 替换 NaN 和 Inf
    df_clean = df.replace([np.inf, -np.inf], None)
    df_clean = df_clean.where(pd.notnull(df_clean), None)
    
    records = []
    for _, row in df_clean.iterrows():
        record = {}
        for col in df_clean.columns:
            record[col] = safe_json_value(row[col])
        records.append(record)
    
    return records

# ===== 配置 =====
SILICONFLOW_API_URL = "https://api.siliconflow.cn/v1/chat/completions"
SILICONFLOW_API_KEY = os.environ.get("SILICONFLOW_API_KEY", "sk-nrryzlftodhwachgnqqmnvwiqnkuhtskjjuezthoofgxprrz")
# 使用更快的模型（QwQ-32B 太慢，换 Qwen2.5-Coder 更适合代码生成）
MODEL_NAME = "Qwen/Qwen2.5-Coder-32B-Instruct"

DEFAULT_PARAMS = {
    "max_tokens": 2048,
    "temperature": 0.3,  # 降低温度，代码生成更稳定
    "top_p": 0.9,
}

# ===== 应用初始化 =====
app = FastAPI(
    title="AI 报表生成器",
    description="基于 QwQ-32B 的智能报表处理 API",
    version="1.0.0",
)

# CORS 配置
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ===== 全局状态 =====
# 存储已加载的表格数据
loaded_tables: Dict[str, Dict[str, Any]] = {}
# 表格元信息（用于校验）
tables_meta: Dict[str, Dict[str, Any]] = {}
# 别名到表名的映射
alias_to_table: Dict[str, str] = {}
# 下一个可用的别名索引
next_alias_index: int = 0
# 步骤执行的中间结果存储
step_results: Dict[str, pd.DataFrame] = {}
# 历史执行记录存储
execution_history: List[Dict[str, Any]] = []
# 历史记录ID计数器
history_id_counter: int = 0


def get_next_alias() -> str:
    """获取下一个可用的别名（A, B, C, ..., Z, AA, AB, ...）"""
    global next_alias_index
    index = next_alias_index
    next_alias_index += 1
    
    result = ""
    while True:
        result = chr(ord('A') + index % 26) + result
        index = index // 26 - 1
        if index < 0:
            break
    return result


def reset_alias_counter():
    """重置别名计数器（当所有表都被删除时）"""
    global next_alias_index
    next_alias_index = 0
    alias_to_table.clear()


# ===== 请求/响应模型 =====
class GenerateRequest(BaseModel):
    """代码生成请求"""
    user_description: str
    selected_fields: Optional[List[Dict[str, str]]] = None


class ExecuteRequest(BaseModel):
    """代码执行请求"""
    code: str


class OperationRequest(BaseModel):
    """操作请求（带校验）"""
    operation: Dict[str, Any]
    force: bool = False  # 是否强制执行（跳过校验）


class ClarificationResponse(BaseModel):
    """用户澄清响应"""
    question_id: str
    answer: Any


# ===== 辅助函数 =====
def simplify_table_name(name: str) -> str:
    """
    简化表名，去掉 UUID 后缀，保留核心名称
    例如: supervision_data_48754d60_83f4_4859_b0d0_e9be20010e3f -> supervision_data
    """
    import re
    # 去掉 UUID 格式的后缀 (8-4-4-4-12 或 下划线分隔的连续hex)
    # 匹配: _xxxxxxxx_xxxx_xxxx_xxxx_xxxxxxxxxxxx 或 _xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx
    uuid_pattern = r'_[0-9a-fA-F]{8}[-_][0-9a-fA-F]{4}[-_][0-9a-fA-F]{4}[-_][0-9a-fA-F]{4}[-_][0-9a-fA-F]{12}$'
    name = re.sub(uuid_pattern, '', name)
    
    # 去掉末尾的纯数字后缀（如 _123456）
    name = re.sub(r'_\d{5,}$', '', name)
    
    # 如果名字为空，给个默认值
    if not name:
        name = "table"
    
    # 替换非法字符为下划线
    name = re.sub(r'[^a-zA-Z0-9_\u4e00-\u9fff]', '_', name)
    
    # 确保不以数字开头（Python 变量名规则）
    if name[0].isdigit():
        name = "t_" + name
    
    return name


def parse_table(file_path: str, sheet_name: Optional[str] = None) -> tuple:
    """
    解析表格文件
    
    Returns:
        (DataFrame, 元信息字典)
    """
    path = Path(file_path)
    
    if not path.exists():
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    suffix = path.suffix.lower()
    sheets_info = []
    
    if suffix == ".csv":
        df = pd.read_csv(file_path, encoding="utf-8-sig")
        sheets_info = ["Sheet1"]
    elif suffix in [".xlsx", ".xls"]:
        # 先获取所有 sheet 名称
        xl = pd.ExcelFile(file_path)
        sheets_info = xl.sheet_names
        
        # 读取指定的 sheet
        if sheet_name and sheet_name in sheets_info:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(file_path)  # 读取第一个 sheet
    else:
        raise ValueError(f"不支持的文件格式: {suffix}")
    
    # 简化表名
    simple_name = simplify_table_name(path.stem)
    
    # 生成元信息（使用安全的 JSON 转换）
    meta = {
        "file_name": path.name,
        "table_name": simple_name,  # 使用简化后的表名
        "original_name": path.stem,  # 保留原始名
        "columns": df.columns.tolist(),
        "dtypes": {col: str(dtype) for col, dtype in df.dtypes.items()},
        "row_count": len(df),
        "sheets": sheets_info,
        "sample": safe_dataframe_to_dict(df.head(3)),
    }
    
    return df, meta


def call_llm(prompt: str, system_prompt: str = None) -> str:
    """调用硅基流动 LLM API"""
    messages = []
    
    if system_prompt:
        messages.append({"role": "system", "content": system_prompt})
    
    messages.append({"role": "user", "content": prompt})
    
    payload = {
        "model": MODEL_NAME,
        "messages": messages,
        "stream": False,
        **DEFAULT_PARAMS,
    }
    
    headers = {
        "Authorization": f"Bearer {SILICONFLOW_API_KEY}",
        "Content-Type": "application/json",
    }
    
    # 创建无代理的 session（解决代理 SSL 问题）
    session = requests.Session()
    session.trust_env = False  # 不使用系统代理
    
    try:
        response = session.post(
            SILICONFLOW_API_URL,
            headers=headers,
            json=payload,
            timeout=120,  # 2分钟超时（新模型更快）
            verify=True,
        )
        response.raise_for_status()
        
        result = response.json()
        return result["choices"][0]["message"]["content"]
    
    except requests.exceptions.RequestException as e:
        return f"API 调用失败: {str(e)}"
    except (KeyError, IndexError) as e:
        return f"解析响应失败: {str(e)}"
    finally:
        session.close()


def extract_code(response: str) -> str:
    """从模型响应中提取 Python 代码"""
    import re
    
    # 尝试提取 ```python ... ``` 包裹的代码
    pattern = r"```python\s*(.*?)\s*```"
    matches = re.findall(pattern, response, re.DOTALL)
    
    if matches:
        return matches[0].strip()
    
    # 尝试提取 ``` ... ``` 包裹的代码
    pattern = r"```\s*(.*?)\s*```"
    matches = re.findall(pattern, response, re.DOTALL)
    
    if matches:
        return matches[0].strip()
    
    return response.strip()


def create_safe_globals(dataframes: Dict[str, pd.DataFrame]) -> Dict[str, Any]:
    """创建安全的执行环境"""
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import uuid
    
    ALLOWED_BUILTINS = {
        'len', 'range', 'enumerate', 'zip', 'map', 'filter',
        'sum', 'min', 'max', 'abs', 'round',
        'list', 'dict', 'set', 'tuple', 'str', 'int', 'float', 'bool',
        'sorted', 'reversed', 'True', 'False', 'None',
    }
    
    safe_builtins = {}
    for k in ALLOWED_BUILTINS:
        if hasattr(__builtins__, k):
            safe_builtins[k] = getattr(__builtins__, k)
        elif isinstance(__builtins__, dict) and k in __builtins__:
            safe_builtins[k] = __builtins__[k]
    
    # 预定义的工作区操作函数
    def add_to_report(df: pd.DataFrame, sheet_name: str, origin: str = "transform") -> pd.DataFrame:
        """
        将 DataFrame 添加到当前报表工作区的新 Sheet
        自动成为活动 Sheet
        """
        ws = workspace_manager.get_or_create_current()
        
        # 找到 df 的来源表名
        source_tables = []
        for var_name, var_df in dataframes.items():
            if var_df is df or (isinstance(var_df, pd.DataFrame) and var_df.equals(df)):
                source_tables.append(var_name.replace('df_', ''))
                break
        
        final_name = ws.add_sheet(sheet_name, df, origin=origin, source_tables=source_tables, set_active=True)
        
        return pd.DataFrame({
            '操作': ['已添加到报表'],
            '当前Sheet': [final_name],
            '报表': [ws.name],
            '说明': [f'Sheet「{final_name}」已添加并设为当前活动表，后续操作将基于此表']
        })
    
    def merge_to_sheet(source_df: pd.DataFrame, target_df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
        """
        将 source_df 合并后添加到报表工作区
        自动成为活动 Sheet
        """
        ws = workspace_manager.get_or_create_current()
        
        # 如果目标表还没在工作区，先添加进去
        target_name = None
        for var_name, var_df in dataframes.items():
            if var_df is target_df or (isinstance(var_df, pd.DataFrame) and var_df.equals(target_df)):
                target_name = var_name.replace('df_', '')
                break
        
        if target_name and target_name not in ws.sheets:
            ws.add_sheet(target_name, target_df, origin="import", set_active=False)
        
        # 添加合并结果
        source_tables = []
        for var_name, var_df in dataframes.items():
            if var_df is source_df or (isinstance(var_df, pd.DataFrame) and var_df.equals(source_df)):
                source_tables.append(var_name.replace('df_', ''))
        if target_name:
            source_tables.append(target_name)
        
        final_name = ws.add_sheet(sheet_name, source_df, origin="merge", source_tables=source_tables, set_active=True)
        
        return pd.DataFrame({
            '操作': ['合并成功'],
            '当前Sheet': [final_name],
            '报表': [ws.name],
            '说明': [f'Sheet「{final_name}」已创建并设为当前活动表，后续操作将基于此表']
        })
    
    def export_multi_sheet(sheets_data: list) -> pd.DataFrame:
        """
        将多个 DataFrame 添加到报表工作区的多个 Sheet
        最后一个 Sheet 成为活动 Sheet
        """
        ws = workspace_manager.get_or_create_current()
        
        for i, (df, sheet_name) in enumerate(sheets_data):
            set_active = (i == len(sheets_data) - 1)  # 最后一个设为活动
            ws.add_sheet(sheet_name, df, origin="import", set_active=set_active)
        
        return pd.DataFrame({
            '操作': ['多表已添加'],
            '当前Sheet': [ws.active_sheet],
            '报表': [ws.name],
            'Sheet列表': [', '.join(ws.sheets.keys())],
            '说明': [f'已添加 {len(sheets_data)} 个 Sheet，当前活动表为「{ws.active_sheet}」']
        })
    
    def get_active_sheet() -> pd.DataFrame:
        """获取当前活动 Sheet 的数据"""
        ws = workspace_manager.get_current()
        if not ws or not ws.active_sheet:
            raise ValueError("没有活动的报表工作区，请先添加数据")
        return ws.get_sheet_df()
    
    safe_globals = {
        '__builtins__': safe_builtins,
        'pd': pd,
        'np': np,
        # 工作区操作函数
        'add_to_report': add_to_report,
        'merge_to_sheet': merge_to_sheet,
        'export_multi_sheet': export_multi_sheet,
        'get_active_sheet': get_active_sheet,
    }
    
    safe_globals.update(dataframes)
    return safe_globals


def execute_code(code: str, dataframes: Dict[str, pd.DataFrame]) -> tuple:
    """
    执行代码
    
    Returns:
        (是否成功, 结果, 错误信息)
    """
    import re as re_module
    
    # 自动移除 import 语句（pandas/numpy 已经预加载）
    code = re_module.sub(r'^import\s+pandas\s+as\s+pd\s*$', '# pandas 已预加载为 pd', code, flags=re_module.MULTILINE)
    code = re_module.sub(r'^import\s+numpy\s+as\s+np\s*$', '# numpy 已预加载为 np', code, flags=re_module.MULTILINE)
    code = re_module.sub(r'^import\s+pandas\s*$', '# pandas 已预加载', code, flags=re_module.MULTILINE)
    code = re_module.sub(r'^import\s+numpy\s*$', '# numpy 已预加载', code, flags=re_module.MULTILINE)
    code = re_module.sub(r'^from\s+pandas\s+import\s+.*$', '# pandas 已预加载为 pd', code, flags=re_module.MULTILINE)
    code = re_module.sub(r'^from\s+numpy\s+import\s+.*$', '# numpy 已预加载为 np', code, flags=re_module.MULTILINE)
    
    # 安全检查
    dangerous_keywords = [
        'import os', 'import sys', 'import subprocess',
        '__import__', 'eval(', 'exec(',
        'open(', 'file(',
        'os.', 'sys.', 'subprocess.',
    ]
    
    # 不允许的文件操作（但允许使用预定义的 merge_to_sheet 和 export_multi_sheet）
    banned_file_ops = [
        'to_excel', 'to_csv', 'to_json', 'to_parquet', 'to_pickle',
        'ExcelWriter', 'read_excel', 'read_csv',
    ]
    
    for keyword in dangerous_keywords:
        if keyword in code:
            return False, None, f"安全检查失败: 代码包含不允许的操作 '{keyword}'"
    
    for keyword in banned_file_ops:
        if keyword in code:
            return False, None, f"⚠️ 请使用预定义函数进行文件操作：\n• merge_to_sheet(源表, 目标表, 'Sheet名') - 合并到新Sheet\n• export_multi_sheet([(表1,'名1'),(表2,'名2')]) - 多表导出"
    
    safe_globals = create_safe_globals(dataframes)
    local_vars = {}
    
    try:
        exec(code, safe_globals, local_vars)
        
        if 'result' in local_vars:
            result = local_vars['result']
            return True, result, ""
        else:
            return False, None, "代码执行完成，但没有找到 'result' 变量"
    
    except Exception as e:
        error_msg = f"{type(e).__name__}: {str(e)}"
        return False, None, error_msg


# ===== API 路由 =====

@app.get("/")
async def root():
    """健康检查"""
    return {"status": "ok", "message": "AI 报表生成器后端运行中"}


@app.get("/api/health")
async def health_check():
    """健康检查"""
    return {"status": "ok"}


@app.post("/api/upload")
async def upload_file(
    file: UploadFile = File(...),
    sheet_name: Optional[str] = Form(None),
    custom_table_name: Optional[str] = Form(None)
):
    """
    上传表格文件
    
    Parameters:
    - file: 上传的文件
    - sheet_name: 指定要加载的 Sheet（可选）
    - custom_table_name: 自定义表名（可选）
    """
    try:
        # 保存文件到临时目录
        temp_dir = Path("./temp")
        temp_dir.mkdir(exist_ok=True)
        
        file_path = temp_dir / file.filename
        
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        # 解析表格（支持指定 sheet）
        df, meta = parse_table(str(file_path), sheet_name=sheet_name)
        
        # 使用自定义表名或默认表名
        base_name = custom_table_name if custom_table_name else meta["table_name"]
        table_name = base_name
        counter = 1
        while table_name in loaded_tables:
            table_name = f"{base_name}_{counter}"
            counter += 1
        meta["table_name"] = table_name
        
        # 获取所有 sheets 信息
        suffix = file_path.suffix.lower()
        if suffix in [".xlsx", ".xls"]:
            xl = pd.ExcelFile(str(file_path))
            all_sheets = xl.sheet_names
        else:
            all_sheets = ["Sheet1"]
        
        meta["sheets"] = all_sheets
        meta["current_sheet"] = sheet_name or all_sheets[0]
        meta["header_row"] = 1
        meta["file_path"] = str(file_path)
        
        # 分配别名
        alias = get_next_alias()
        alias_to_table[alias] = table_name
        
        # 存储到全局状态
        loaded_tables[table_name] = {
            "df": df,
            "path": str(file_path),
            "meta": meta,
            "alias": alias,           # 系统分配的别名 (A, B, C...)
            "nickname": "",           # 用户自定义的昵称
        }
        tables_meta[table_name] = meta
        
        return {
            "success": True,
            "table_name": table_name,
            "meta": meta,
            "table": {
                "name": table_name,
                "alias": alias,
                "nickname": "",
                "columns": meta["columns"],
                "row_count": meta["row_count"],
                "sheets": all_sheets,
                "current_sheet": meta["current_sheet"],
                "header_row": 1,
                "file_path": str(file_path),
            }
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


class ReloadTableRequest(BaseModel):
    table_name: str
    sheet_name: Optional[str] = None
    header_row: int = 1  # 1-based，第几行是表头


@app.post("/api/reload-table")
async def reload_table(request: ReloadTableRequest):
    """
    使用新的设置重新加载表格（切换 Sheet、调整表头行）
    """
    if request.table_name not in loaded_tables:
        raise HTTPException(status_code=404, detail=f"表格 '{request.table_name}' 不存在")
    
    try:
        table_info = loaded_tables[request.table_name]
        file_path = table_info["path"]
        
        suffix = Path(file_path).suffix.lower()
        
        # header_row 转为 0-based 索引（pandas 用 header 参数）
        header_idx = request.header_row - 1 if request.header_row > 0 else 0
        
        if suffix in [".xlsx", ".xls"]:
            xl = pd.ExcelFile(file_path)
            sheet_name = request.sheet_name or xl.sheet_names[0]
            
            # 读取数据，指定表头行
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_idx)
            sheets = xl.sheet_names
        else:
            df = pd.read_csv(file_path, encoding="utf-8-sig", header=header_idx)
            sheet_name = "Sheet1"
            sheets = ["Sheet1"]
        
        # 清理列名
        df.columns = [str(col).strip() if pd.notna(col) else f"列{i+1}" for i, col in enumerate(df.columns)]
        
        # 更新元信息
        meta = {
            "table_name": request.table_name,
            "columns": df.columns.tolist(),
            "dtypes": {col: str(df[col].dtype) for col in df.columns},
            "row_count": len(df),
            "sample": safe_dataframe_to_dict(df.head(5)),
            "sheets": sheets,
            "current_sheet": sheet_name,
            "header_row": request.header_row,
        }
        
        # 更新全局状态
        loaded_tables[request.table_name]["df"] = df
        loaded_tables[request.table_name]["meta"] = meta
        tables_meta[request.table_name] = meta
        
        return {
            "success": True,
            "message": f"已重新加载表格，使用 Sheet '{sheet_name}'，表头行 {request.header_row}",
            "table": {
                "name": request.table_name,
                "columns": meta["columns"],
                "row_count": meta["row_count"],
                "sheets": sheets,
                "current_sheet": sheet_name,
                "header_row": request.header_row,
            }
        }
    
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class PreviewSheetRequest(BaseModel):
    table_name: str
    sheet_name: str
    limit: int = 20


class SplitSheetRequest(BaseModel):
    split_id: str
    sheet_name: str
    limit: int = 20


@app.post("/api/split-sheet-data")
async def get_split_sheet_data(request: SplitSheetRequest):
    """
    获取拆分结果中某个 Sheet 的数据
    用于前端 Sheet 标签切换
    """
    if request.split_id not in split_results:
        raise HTTPException(status_code=404, detail="拆分结果不存在或已过期")
    
    split_data = split_results[request.split_id]
    
    if request.sheet_name not in split_data["sheets"]:
        raise HTTPException(status_code=404, detail=f"Sheet '{request.sheet_name}' 不存在")
    
    sheet_info = split_data["sheets"][request.sheet_name]
    df = sheet_info["df"]
    
    return {
        "columns": df.columns.tolist(),
        "data": safe_dataframe_to_dict(df.head(request.limit)),
        "total_rows": len(df),
        "sheet_name": request.sheet_name,
    }


@app.post("/api/preview-sheet")
async def preview_sheet(request: PreviewSheetRequest):
    """
    预览指定表格的某个 Sheet 数据（不修改当前加载状态）
    用于前端 Sheet 标签切换时快速预览
    """
    if request.table_name not in loaded_tables:
        raise HTTPException(status_code=404, detail=f"表格 '{request.table_name}' 不存在")
    
    try:
        table_info = loaded_tables[request.table_name]
        file_path = table_info.get("path")
        
        if not file_path or not Path(file_path).exists():
            raise HTTPException(status_code=404, detail="原始文件不存在")
        
        suffix = Path(file_path).suffix.lower()
        
        if suffix in [".xlsx", ".xls"]:
            # 读取指定 Sheet
            df = pd.read_excel(file_path, sheet_name=request.sheet_name)
        else:
            # CSV 只有一个 Sheet
            df = pd.read_csv(file_path, encoding="utf-8-sig")
        
        # 清理列名
        df.columns = [str(col).strip() if pd.notna(col) else f"列{i+1}" for i, col in enumerate(df.columns)]
        
        # 返回预览数据
        return {
            "columns": df.columns.tolist(),
            "data": safe_dataframe_to_dict(df.head(request.limit)),
            "total_rows": len(df),
            "sheet_name": request.sheet_name,
        }
    
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/upload-preview")
async def upload_preview(file: UploadFile = File(...)):
    """
    预览上传文件的 Sheets 信息（不实际加载数据）
    """
    try:
        temp_dir = Path("./temp")
        temp_dir.mkdir(exist_ok=True)
        
        file_path = temp_dir / file.filename
        
        with open(file_path, "wb") as f:
            content = await file.read()
            f.write(content)
        
        suffix = file_path.suffix.lower()
        sheets_info = []
        
        if suffix in [".xlsx", ".xls"]:
            xl = pd.ExcelFile(str(file_path))
            for sheet in xl.sheet_names:
                # 读取每个 sheet 的行数
                df_temp = pd.read_excel(str(file_path), sheet_name=sheet, nrows=0)
                df_full = pd.read_excel(str(file_path), sheet_name=sheet)
                sheets_info.append({
                    "name": sheet,
                    "columns": df_temp.columns.tolist(),
                    "row_count": len(df_full),
                })
        else:
            df = pd.read_csv(str(file_path), encoding="utf-8-sig")
            sheets_info.append({
                "name": "Sheet1",
                "columns": df.columns.tolist(),
                "row_count": len(df),
            })
        
        return {
            "success": True,
            "file_name": file.filename,
            "file_path": str(file_path),
            "sheets": sheets_info,
        }
    
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/tables")
async def list_tables():
    """
    列出已加载的表格
    """
    result = []
    for name, info in loaded_tables.items():
        result.append({
            "name": name,
            "alias": info.get("alias", ""),
            "nickname": info.get("nickname", ""),
            "columns": info["meta"]["columns"],
            "row_count": info["meta"]["row_count"],
            "sheets": info["meta"].get("sheets", []),
            "current_sheet": info["meta"].get("current_sheet", ""),
            "header_row": info["meta"].get("header_row", 1),
        })
    
    # 按别名排序
    result.sort(key=lambda x: x.get("alias", ""))
    
    return {"tables": result}


class UpdateNicknameRequest(BaseModel):
    table_name: str
    nickname: str


@app.post("/api/update-nickname")
async def update_nickname(request: UpdateNicknameRequest):
    """
    更新表的昵称
    """
    if request.table_name not in loaded_tables:
        raise HTTPException(status_code=404, detail=f"表格 '{request.table_name}' 不存在")
    
    loaded_tables[request.table_name]["nickname"] = request.nickname
    
    return {
        "success": True,
        "table_name": request.table_name,
        "nickname": request.nickname,
    }


@app.delete("/api/tables/{table_name}")
async def delete_table(table_name: str):
    """
    删除表格
    """
    if table_name not in loaded_tables:
        raise HTTPException(status_code=404, detail=f"表格 '{table_name}' 不存在")
    
    # 获取别名并从映射中删除
    alias = loaded_tables[table_name].get("alias", "")
    if alias in alias_to_table:
        del alias_to_table[alias]
    
    # 删除表格
    del loaded_tables[table_name]
    if table_name in tables_meta:
        del tables_meta[table_name]
    
    # 如果所有表都删除了，重置别名计数器
    if not loaded_tables:
        reset_alias_counter()
    
    return {"success": True, "message": f"已删除表格 '{table_name}'"}


@app.delete("/api/tables")
async def clear_all_tables():
    """
    清空所有表格并重置别名计数器
    """
    loaded_tables.clear()
    tables_meta.clear()
    reset_alias_counter()
    
    return {"success": True, "message": "已清空所有表格"}


@app.post("/api/generate")
async def generate_code(request: GenerateRequest):
    """
    根据自然语言生成 Pandas 代码
    """
    if not loaded_tables:
        raise HTTPException(status_code=400, detail="请先上传表格文件")
    
    # 构建表格信息
    table_info = {}
    for name, info in loaded_tables.items():
        meta = info["meta"]
        alias = info.get("alias", name)
        nickname = info.get("nickname", "")
        table_info[name] = {
            "alias": alias,
            "nickname": nickname,
            "var_name": f"df_{alias}",  # 使用别名作为变量名
            "display_name": nickname if nickname else name,  # 显示名称
            "columns": meta["columns"],
            "dtypes": meta["dtypes"],
            "sample": meta["sample"],
        }
    
    # 构建系统提示词
    system_prompt = """你是一个专业的 pandas 数据分析代码生成器。

【重要规则】
1. 只输出可直接执行的 Python 代码，不要任何解释文字
2. 代码用 ```python 和 ``` 包裹
3. ⚠️ 不要写 import 语句！pd 和 np 已经预加载
4. ⚠️ 变量名规则：每个表对应一个 DataFrame 变量，变量名格式为 df_表名
5. ⚠️ 必须使用下方列出的【精确变量名】，不要自己编造！
6. 最终结果必须存入名为 `result` 的变量
7. 不要使用 print、to_excel、to_csv 等
8. ⚠️ 不要使用 .plot()！如需图表，返回统计数据 DataFrame

【⚠️ 关键：数据来源选择】
- ✅ 默认使用 df_表名 变量！这是主要的数据来源！
- ❌ 不要使用 get_active_sheet()，除非用户明确说"当前表"、"刚才的表"、"继续处理"
- 如果用户说"统计xxx"、"筛选xxx"，直接使用 df_表名

【⚠️ 特殊指令：拆分导出】
如果用户要求"按某字段拆分成多个sheet导出"或类似请求，不要生成代码！
直接输出这个格式（不要代码块）：
@SPLIT_EXPORT:表名:列名

例如用户说"按部门拆分A表导出"：
@SPLIT_EXPORT:原始表名:部门

注意：表名使用原始表名（不是df_前缀的变量名），列名使用精确的列名

【预定义函数】（仅在需要时使用）
- add_to_report(df, "Sheet名") → 将结果添加到报表的新Sheet
- merge_to_sheet(源df, 目标df, "Sheet名") → 合并两表到新Sheet

【示例】
用户说"把表A放到表B的sheet2里，命名为xxx"：
result = merge_to_sheet(df_表A, df_表B, "xxx")

用户说"统计各状态数量"：
result = df_表A['状态'].value_counts().reset_index()
result.columns = ['状态', '数量']

用户说"筛选金额大于1000的记录"：
result = df_表A[df_表A['金额'] > 1000]

用户说"vlookup/关联：根据表A的xxx列和表B的yyy列，把表B的zzz列关联到表A"：
result = pd.merge(df_表A, df_表B[['yyy', 'zzz']], left_on='xxx', right_on='yyy', how='left')
# 如果关联键名称不同，保留原表A的列，可以删除多余的关联键列
result = result.drop(columns=['yyy'], errors='ignore')

用户说"把表B的编号匹配到表A（按名称匹配）"：
result = pd.merge(df_表A, df_表B[['名称', '编号']], left_on='表A的名称列', right_on='名称', how='left').drop(columns=['名称'], errors='ignore')"""

    # 构建表格描述
    tables_desc = []
    for table_name, info in table_info.items():
        alias = info["alias"]
        var_name = info["var_name"]
        nickname = info["nickname"]
        display_name = info["display_name"]
        
        cols_desc = [f"  - {col} ({info['dtypes'].get(col, 'unknown')})" 
                     for col in info["columns"]]
        
        sample_df = pd.DataFrame(info["sample"])
        sample_str = f"\n示例数据:\n{sample_df.to_string(index=False)}"
        
        # 构建用户可能使用的引用方式
        ref_names = [f"{alias}表", f"表{alias}"]
        if nickname:
            ref_names.append(nickname)
        ref_names.append(table_name)
        
        tables_desc.append(f"""### {alias}表: {display_name}
变量名: `{var_name}` ← 必须使用这个变量名！
用户可能的引用方式: {', '.join(ref_names)}
原表名: {table_name}
列信息:
{chr(10).join(cols_desc)}
{sample_str}""")
    
    # 构建字段选择描述
    fields_desc = []
    if request.selected_fields:
        for field in request.selected_fields:
            fields_desc.append(f"- {field['table']}.{field['column']} → {field.get('role', '可用')}")
    
    # 构建完整提示词
    prompt = f"""## 可用数据表

{chr(10).join(tables_desc)}

## 用户选择的字段
{chr(10).join(fields_desc) if fields_desc else "（未指定，请根据需求自行选择）"}

## 用户需求描述
{request.user_description}

【重要】请严格按以下格式回复：

【分析】在这里用1-2句话简要说明：1)你理解用户想做什么 2)你打算用什么方法实现

```python
# 在这里写代码
result = ...
```

【示例】
用户说"按区域统计销售额"，你应该回复：

【分析】用户需要按区域汇总销售额。我将使用 groupby 对区域分组并求和销售额。

```python
result = df_销售数据.groupby('区域')['销售额'].sum().reset_index()
```"""

    # 调用模型
    response = call_llm(prompt, system_prompt)
    code = extract_code(response)
    
    # 提取分析过程（多种格式兼容）
    import re
    analysis = ""
    
    # 尝试多种匹配模式
    patterns = [
        r'【分析】(.+?)(?=```|$)',           # 【分析】格式
        r'\[分析\](.+?)(?=```|$)',           # [分析]格式
        r'分析[：:]\s*(.+?)(?=```|$)',       # 分析：格式
        r'^(.+?)(?=```)',                    # 代码块前的所有文字
    ]
    
    for pattern in patterns:
        match = re.search(pattern, response, re.DOTALL)
        if match:
            text = match.group(1).strip()
            # 过滤掉太短或无意义的内容
            if len(text) > 10 and not text.startswith('python'):
                analysis = text
                break
    
    # 清理分析文本
    if analysis:
        # 移除可能的标签残留
        analysis = re.sub(r'^[【\[](分析|思路|理解)[】\]][：:]?\s*', '', analysis)
        analysis = analysis.strip()
    
    print(f"[Debug] Analysis extracted: {analysis[:100] if analysis else 'None'}...")
    
    return {
        "success": True,
        "code": code,
        "analysis": analysis,
        "raw_response": response,
    }


@app.post("/api/execute")
async def execute(request: ExecuteRequest):
    """
    执行生成的代码
    """
    if not loaded_tables:
        raise HTTPException(status_code=400, detail="请先上传表格文件")
    
    # 准备 DataFrame 变量（使用别名）
    dataframes = {}
    for name, info in loaded_tables.items():
        alias = info.get("alias", name)
        dataframes[f"df_{alias}"] = info["df"]
        # 也保留原表名作为变量（兼容性）
        dataframes[f"df_{name}"] = info["df"]
    
    # 执行代码
    success, result, error = execute_code(request.code, dataframes)
    
    if not success:
        return {
            "success": False,
            "error": error,
        }
    
    if result is None:
        return {
            "success": False,
            "error": error or "执行完成但没有结果",
        }
    
    # 格式化结果
    if isinstance(result, pd.DataFrame):
        return {
            "success": True,
            "result": {
                "type": "dataframe",
                "columns": result.columns.tolist(),
                "data": safe_dataframe_to_dict(result.head(100)),
                "total_rows": len(result),
            },
        }
    elif isinstance(result, pd.Series):
        # Series 也需要安全转换
        series_dict = {}
        for k, v in result.items():
            series_dict[str(k)] = safe_json_value(v)
        return {
            "success": True,
            "result": {
                "type": "series",
                "data": series_dict,
            },
        }
    else:
        return {
            "success": True,
            "result": {
                "type": "other",
                "data": str(result),
            },
        }


@app.post("/api/validate")
async def validate_operation(request: OperationRequest):
    """
    校验操作意图（在执行前）
    """
    if request.force:
        return {"status": "valid", "message": "强制执行模式，跳过校验"}
    
    validator = OperationValidator(tables_meta)
    result = validator.validate_operation(request.operation)
    
    # 格式化消息
    if result["status"] == "need_clarification":
        result["formatted_message"] = format_clarification_message(result)
    
    return result


def detect_chart_intent(query: str) -> dict:
    """
    检测用户是否需要图表，以及需要什么类型的图表
    """
    query_lower = query.lower()
    
    # 饼图关键词
    pie_keywords = ['饼图', '饼状图', '占比图', '比例图', 'pie']
    # 柱状图关键词
    bar_keywords = ['柱状图', '柱形图', '条形图', 'bar', '直方图']
    # 通用图表关键词
    chart_keywords = ['统计图', '图表', '可视化', '图形', 'chart', '趋势图', '分布图']
    
    for kw in pie_keywords:
        if kw in query_lower:
            return {"show_chart": True, "chart_type": "pie"}
    
    for kw in bar_keywords:
        if kw in query_lower:
            return {"show_chart": True, "chart_type": "bar"}
    
    for kw in chart_keywords:
        if kw in query_lower:
            return {"show_chart": True, "chart_type": "bar"}  # 默认柱状图
    
    # 没有明确要求图表
    return {"show_chart": False, "chart_type": None}


# 存储拆分结果（用于前端 Sheet 切换）
split_results = {}


async def handle_split_export(user_desc: str, split_info: dict):
    """处理拆分请求 - 在页面中显示多个 Sheet 标签"""
    import uuid
    
    table_alias = split_info.get("table_alias")
    column_hint = split_info.get("column")
    
    # 找到目标表
    target_table = None
    target_name = None
    
    if table_alias:
        for name, info in loaded_tables.items():
            if info.get("alias") == table_alias:
                target_table = info
                target_name = name
                break
    
    if not target_table:
        if len(loaded_tables) == 1:
            target_name = list(loaded_tables.keys())[0]
            target_table = loaded_tables[target_name]
        else:
            table_options = [f"{info.get('alias', '')}表: {name}" for name, info in loaded_tables.items()]
            return {
                "status": "need_clarification",
                "type": "select_table",
                "message": "请指定要拆分的表格",
                "options": table_options,
            }
    
    df = target_table["df"]
    alias = target_table.get("alias", "")
    
    # 找到拆分列
    split_column = None
    
    if column_hint:
        if column_hint in df.columns:
            split_column = column_hint
        else:
            for col in df.columns:
                if column_hint.lower() == col.lower():
                    split_column = col
                    break
                elif column_hint.lower() in col.lower() or col.lower() in column_hint.lower():
                    split_column = col
                    break
    
    if not split_column:
        columns_info = []
        for col in df.columns:
            unique_count = df[col].nunique()
            if 1 < unique_count <= 100:
                columns_info.append(f"{col} ({unique_count}个不同值)")
        
        return {
            "status": "need_clarification",
            "type": "select_column",
            "message": f"请指定按哪个字段拆分 {alias}表",
            "options": columns_info if columns_info else df.columns.tolist(),
        }
    
    # 检查唯一值数量
    unique_values = df[split_column].dropna().unique()
    
    if len(unique_values) == 0:
        return {"status": "error", "message": f"列 '{split_column}' 没有有效值"}
    
    if len(unique_values) > 50:
        return {"status": "error", "message": f"拆分值过多 ({len(unique_values)} 个)，最多支持 50 个分组"}
    
    # 执行拆分，存储到内存
    try:
        split_id = str(uuid.uuid4())[:8]
        sheets_data = {}
        sheet_names = []
        
        for value in unique_values:
            sub_df = df[df[split_column] == value]
            
            # 清理名称
            sheet_name = str(value)[:31]
            for char in ['\\', '/', '*', '?', ':', '[', ']']:
                sheet_name = sheet_name.replace(char, '_')
            
            original_name = sheet_name
            counter = 1
            while sheet_name in sheet_names:
                sheet_name = f"{original_name[:28]}_{counter}"
                counter += 1
            
            sheet_names.append(sheet_name)
            sheets_data[sheet_name] = {
                "df": sub_df,
                "row_count": len(sub_df),
                "value": str(value),
            }
        
        # 存储拆分结果
        split_results[split_id] = {
            "source_table": target_name,
            "alias": alias,
            "split_column": split_column,
            "sheets": sheets_data,
            "sheet_names": sheet_names,
        }
        
        # 返回第一个 Sheet 的预览数据
        first_sheet = sheet_names[0]
        first_df = sheets_data[first_sheet]["df"]
        
        # 构建 Sheet 信息
        sheets_info = [
            {"name": name, "row_count": sheets_data[name]["row_count"]}
            for name in sheet_names
        ]
        
        return {
            "status": "success",
            "analysis": f"按 **{split_column}** 字段将 {alias}表 拆分成 **{len(sheet_names)}** 个分组",
            "result": {
                "type": "dataframe",
                "columns": first_df.columns.tolist(),
                "data": safe_dataframe_to_dict(first_df.head(20)),
                "total_rows": len(first_df),
            },
            "code": f"# 按 {split_column} 拆分\n# 共 {len(sheet_names)} 个分组: {', '.join(sheet_names[:5])}{'...' if len(sheet_names) > 5 else ''}",
            "chart": {"show_chart": False, "chart_type": None},
            # 拆分结果信息（前端用于显示 Sheet 标签）
            "splitResult": {
                "id": split_id,
                "sheets": sheets_info,
                "currentSheet": first_sheet,
                "splitColumn": split_column,
                "sourceAlias": alias,
            },
        }
        
    except Exception as e:
        traceback.print_exc()
        return {"status": "error", "message": f"拆分失败: {str(e)}"}


def detect_split_export(user_desc: str) -> dict:
    """
    检测是否是拆分导出请求
    返回: {"is_split": bool, "table_alias": str, "column": str}
    """
    import re
    
    # 关键词模式
    split_keywords = ['拆分', '分成多个sheet', '按.*导出', '拆成.*sheet', '分割成.*sheet']
    
    # 检查是否包含拆分关键词
    is_split = False
    for kw in split_keywords:
        if re.search(kw, user_desc, re.IGNORECASE):
            is_split = True
            break
    
    if not is_split:
        return {"is_split": False}
    
    # 提取表名（A表、B表 等）
    table_match = re.search(r'([A-Z])表', user_desc)
    table_alias = table_match.group(1) if table_match else None
    
    # 提取列名（按xxx字段/按xxx拆分）
    column_patterns = [
        r'按[「「]?(\w+)[」」]?[字段列]?拆分',
        r'按[「「]?(\w+)[」」]?[字段列]?分[成割]',
        r'以[「「]?(\w+)[」」]?[字段列]?拆分',
        r'根据[「「]?(\w+)[」」]?[字段列]?拆分',
        r'用[「「]?(\w+)[」」]?[字段列]?拆分',
        r'[「「]?(\w+)[」」]?[字段列]?命名',
        r'[「「]?(\w+)[」」]?[字段列]?为sheet',
    ]
    
    column = None
    for pattern in column_patterns:
        match = re.search(pattern, user_desc, re.IGNORECASE)
        if match:
            column = match.group(1)
            break
    
    # 如果没找到中文列名，尝试英文
    if not column:
        # 提取可能的英文列名
        eng_match = re.search(r'(?:按|by|用)\s*[`「」]?([A-Za-z_][A-Za-z0-9_]*)[`「」]?', user_desc, re.IGNORECASE)
        if eng_match:
            column = eng_match.group(1)
    
    return {
        "is_split": True,
        "table_alias": table_alias,
        "column": column
    }


@app.post("/api/smart-execute")
async def smart_execute(request: GenerateRequest):
    """
    智能执行：生成 → 校验 → 执行（或澄清）
    """
    if not loaded_tables:
        return get_clarification_response(
            "no_tables",
            message="请先上传表格文件",
        )
    
    # 0.0 检测是否是"历史执行处理"指令
    user_desc_lower = request.user_description.lower().strip()
    history_keywords = ["历史执行处理", "历史执行", "执行历史", "重复执行", "批量执行历史"]
    is_history_command = any(kw in user_desc_lower for kw in history_keywords)
    
    if is_history_command:
        # 执行历史命令
        result = await batch_execute_history()
        
        if result.get("status") == "structure_mismatch":
            return {
                "status": "need_clarification",
                "type": "structure_mismatch",
                "message": result["message"],
                "analysis": f"⚠️ 表结构不匹配\n\n{result.get('suggestion', '')}",
                "code": "# 无法执行：表结构不匹配",
                "chart": {"show_chart": False, "chart_type": None},
            }
        
        if result.get("status") == "error":
            return {
                "status": "error",
                "message": result["message"],
                "analysis": "历史执行失败",
                "code": "# 无历史记录",
                "chart": {"show_chart": False, "chart_type": None},
            }
        
        # 构建执行日志显示
        logs = result.get("execution_logs", [])
        log_text = "### 历史执行日志\n\n"
        for log in logs:
            status_icon = "✅" if log["status"] == "success" else "❌"
            log_text += f"{status_icon} **步骤 {log['step']}**: {log['query']}\n"
            if log.get("result_rows"):
                log_text += f"   结果行数: {log['result_rows']}\n"
            if log["status"] == "error":
                log_text += f"   错误: {log.get('message', '未知错误')}\n"
            log_text += "\n"
        
        return {
            "status": "success",
            "analysis": log_text,
            "result": result.get("result"),
            "code": result.get("code", "# 历史批量执行"),
            "chart": {"show_chart": False, "chart_type": None},
            "historyExecution": {
                "total_steps": result.get("total_steps", 0),
                "success_count": result.get("success_count", 0),
                "error_count": result.get("error_count", 0),
                "logs": logs,
            },
        }
    
    # 0. 检测是否是拆分导出请求
    split_info = detect_split_export(request.user_description)
    if split_info["is_split"]:
        return await handle_split_export(request.user_description, split_info)
    
    # 检测图表意图
    chart_intent = detect_chart_intent(request.user_description)
    
    # 1. 生成代码
    analysis = ""
    try:
        gen_result = await generate_code(request)
        code = gen_result["code"]
        analysis = gen_result.get("analysis", "")
    except Exception as e:
        return get_clarification_response(
            "generation_failed",
            message=f"代码生成失败: {str(e)}",
        )
    
    # 1.5 检测是否是拆分导出指令
    import re
    split_match = re.search(r'@SPLIT_EXPORT:([^:]+):(.+)', code)
    if split_match:
        table_name = split_match.group(1).strip()
        split_column = split_match.group(2).strip()
        
        # 查找实际的表名（可能用户说的是别名）
        actual_table = None
        for name, info in loaded_tables.items():
            alias = info.get("alias", "")
            if name == table_name or alias == table_name or f"{alias}表" == table_name:
                actual_table = name
                break
        
        if not actual_table:
            return {
                "status": "error",
                "message": f"找不到表 '{table_name}'",
                "code": code,
            }
        
        # 检查列是否存在
        df = loaded_tables[actual_table]["df"]
        if split_column not in df.columns:
            # 尝试模糊匹配
            matched_col = None
            for col in df.columns:
                if split_column.lower() in col.lower() or col.lower() in split_column.lower():
                    matched_col = col
                    break
            if matched_col:
                split_column = matched_col
            else:
                return {
                    "status": "need_clarification",
                    "type": "missing_column",
                    "message": f"列 '{split_column}' 不存在，请选择正确的列名",
                    "options": df.columns.tolist(),
                    "code": code,
                }
        
        # 执行拆分导出
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            import uuid
            
            unique_values = df[split_column].dropna().unique()
            
            if len(unique_values) == 0:
                return {
                    "status": "error",
                    "message": f"列 '{split_column}' 没有有效值",
                    "code": code,
                }
            
            if len(unique_values) > 100:
                return {
                    "status": "error",
                    "message": f"拆分值过多 ({len(unique_values)} 个)，最多支持 100 个 Sheet",
                    "code": code,
                }
            
            # 创建工作簿
            wb = Workbook()
            default_sheet = wb.active
            sheet_names = []
            
            for i, value in enumerate(unique_values):
                sub_df = df[df[split_column] == value]
                
                # 清理 Sheet 名称
                sheet_name = str(value)[:31]
                for char in ['\\', '/', '*', '?', ':', '[', ']']:
                    sheet_name = sheet_name.replace(char, '_')
                
                original_name = sheet_name
                counter = 1
                while sheet_name in sheet_names:
                    sheet_name = f"{original_name[:28]}_{counter}"
                    counter += 1
                sheet_names.append(sheet_name)
                
                if i == 0:
                    ws = default_sheet
                    ws.title = sheet_name
                else:
                    ws = wb.create_sheet(title=sheet_name)
                
                for r_idx, row in enumerate(dataframe_to_rows(sub_df, index=False, header=True), 1):
                    for c_idx, col_value in enumerate(row, 1):
                        if isinstance(col_value, float) and (np.isnan(col_value) or np.isinf(col_value)):
                            col_value = None
                        ws.cell(row=r_idx, column=c_idx, value=col_value)
            
            # 保存
            temp_dir = Path("./temp/exports")
            temp_dir.mkdir(parents=True, exist_ok=True)
            
            alias = loaded_tables[actual_table].get("alias", "")
            file_id = str(uuid.uuid4())[:8]
            output_name = f"{alias}表_按{split_column}拆分_{file_id}.xlsx"
            output_path = temp_dir / output_name
            
            wb.save(str(output_path))
            
            # 返回成功结果
            result_df = pd.DataFrame([{
                '操作': '拆分导出成功',
                '拆分字段': split_column,
                'Sheet数量': len(sheet_names),
                'Sheet列表': ', '.join(sheet_names[:10]) + ('...' if len(sheet_names) > 10 else ''),
                '文件名': output_name,
                '下载链接': f"/api/download/{output_name}",
            }])
            
            return {
                "status": "success",
                "analysis": f"按 {split_column} 字段将表格拆分成 {len(sheet_names)} 个 Sheet",
                "result": {
                    "type": "dataframe",
                    "columns": result_df.columns.tolist(),
                    "data": safe_dataframe_to_dict(result_df),
                    "total_rows": 1,
                },
                "code": f"# 拆分导出: {actual_table} → 按 {split_column} 拆分成 {len(sheet_names)} 个 Sheet",
                "chart": {"show_chart": False, "chart_type": None},
            }
            
        except Exception as e:
            traceback.print_exc()
            return {
                "status": "error",
                "message": f"拆分导出失败: {str(e)}",
                "code": code,
            }
    
    # 2. 执行代码（带错误处理）
    # 使用别名作为变量名，同时保留原表名兼容
    dataframes = {}
    for name, info in loaded_tables.items():
        alias = info.get("alias", name)
        dataframes[f"df_{alias}"] = info["df"]
        dataframes[f"df_{name}"] = info["df"]
    success, result, error = execute_code(code, dataframes)
    
    if not success:
        # 分析错误类型，返回友好的澄清消息
        if "KeyError" in error:
            # 提取 KeyError 的键名
            import re
            match = re.search(r"KeyError: ['\"]?([^'\"]+)['\"]?", error)
            key_name = match.group(1) if match else "未知"
            
            # 判断是列名还是表名
            all_columns = []
            for meta in tables_meta.values():
                all_columns.extend(meta.get("columns", []))
            
            return {
                "status": "need_clarification",
                "type": "missing_column",
                "message": f"在表中没有找到「{key_name}」",
                "options": list(set(all_columns)),
                "code": code,
                "original_error": error,
            }
        
        return {
            "status": "error",
            "message": error,
            "code": code,
        }
    
    # 3. 返回成功结果（包含图表意图和分析过程）
    
    # 保存到历史记录
    save_to_history(request.user_description, code, analysis)
    
    # 检测代码使用的主要数据源，附加 Sheet 信息
    source_table_info = None
    import re
    df_match = re.search(r'result\s*=\s*df_([A-Z])\b', code)
    if df_match:
        alias = df_match.group(1)
        # 找到对应的表
        for name, info in loaded_tables.items():
            if info.get("alias") == alias:
                meta = info.get("meta", {})
                sheets = meta.get("sheets", [])
                if len(sheets) > 1:
                    source_table_info = {
                        "name": name,
                        "alias": alias,
                        "sheets": sheets,
                        "current_sheet": meta.get("current_sheet"),
                    }
                break
    
    if isinstance(result, pd.DataFrame):
        response = {
            "status": "success",
            "analysis": analysis,  # 分析过程
            "result": {
                "type": "dataframe",
                "columns": result.columns.tolist(),
                "data": safe_dataframe_to_dict(result.head(100)),
                "total_rows": len(result),
            },
            "code": code,
            "chart": chart_intent,  # 图表意图
        }
        if source_table_info:
            response["sourceTable"] = source_table_info
        return response
    elif isinstance(result, pd.Series):
        # Series 安全转换
        series_dict = {}
        for k, v in result.items():
            series_dict[str(k)] = safe_json_value(v)
        return {
            "status": "success",
            "analysis": analysis,  # 分析过程
            "result": {
                "type": "series",
                "data": series_dict,
            },
            "code": code,
            "chart": chart_intent,  # 图表意图
        }
    else:
        return {
            "status": "success",
            "analysis": analysis,  # 分析过程
            "result": {
                "type": "other",
                "data": str(result),
            },
            "code": code,
            "chart": {"show_chart": False, "chart_type": None},
        }


# ===== 任务分解与步骤执行 =====

class TaskStep(BaseModel):
    """单个执行步骤"""
    id: int
    description: str
    input_tables: List[str]  # 输入来源（表别名或步骤结果ID）
    output_id: str           # 输出结果的 ID
    code: Optional[str] = None  # 生成的代码（确认后填充）


class TaskPlan(BaseModel):
    """任务执行计划"""
    is_multi_step: bool
    steps: List[TaskStep]
    original_query: str


class AnalyzeTaskRequest(BaseModel):
    """任务分析请求"""
    user_description: str


class ExecuteStepRequest(BaseModel):
    """单步执行请求"""
    step_id: int
    step_description: str
    input_sources: List[str]  # 输入来源（别名或 step_X 格式）
    output_id: str


def is_multi_step_query(query: str) -> bool:
    """
    判断是否是多步骤任务
    基于关键词和语句结构判断
    """
    import re
    query_lower = query.lower()
    
    # ===== 单步骤模式（优先识别，直接返回 False）=====
    # vlookup/匹配/关联操作 → 绝对单步（即使有"然后"也是单步）
    # 因为 vlookup 本质上就是一个 merge 操作
    vlookup_patterns = [
        r'vlookup',
        r'lookup',
        r'匹配.*到.*表',
        r'关联.*到.*表',
        r'(?:把|将).*(?:匹配|关联|vlookup).*到',
    ]
    
    for pattern in vlookup_patterns:
        if re.search(pattern, query_lower):
            # vlookup 类型操作，强制单步执行
            return False
    
    # ===== 多步骤指示词 =====
    # 只有真正的多步骤流程才分解
    multi_step_indicators = [
        r'先.*然后.*最后',      # 明确的三步流程
        r'第一步.*第二步',      # 明确的步骤编号
        r'首先.*其次.*然后',    # 顺序词
        r'分别.*汇总.*合并',    # 分别处理再合并
    ]
    
    # 检查多步骤指示词
    for indicator in multi_step_indicators:
        if re.search(indicator, query_lower):
            return True
    
    # 检查是否涉及多个表 + 复杂聚合操作
    table_refs = set()
    for alias in alias_to_table.keys():
        if f'{alias}表' in query or f'表{alias}' in query:
            table_refs.add(alias)
    
    # 涉及多个表 + 有复杂聚合操作 → 多步骤
    if len(table_refs) >= 3:
        complex_ops = ['汇总', '统计', '分组', '求和', '计数', '平均', '透视']
        has_complex = any(op in query for op in complex_ops)
        if has_complex:
            return True
    
    return False


@app.post("/api/analyze-task")
async def analyze_task(request: AnalyzeTaskRequest):
    """
    分析用户任务，判断是否需要多步骤执行，并生成执行计划
    只分析，不执行
    """
    if not loaded_tables:
        raise HTTPException(status_code=400, detail="请先上传表格文件")
    
    query = request.user_description
    
    # 判断是否是多步骤任务
    if not is_multi_step_query(query):
        # 单步任务，不需要计划确认
        return {
            "is_multi_step": False,
            "message": "这是一个简单任务，可以直接执行",
        }
    
    # 构建表格信息用于 LLM 分析
    table_info = []
    for name, info in loaded_tables.items():
        alias = info.get("alias", name)
        nickname = info.get("nickname", "")
        display_name = nickname if nickname else name
        columns = info["meta"]["columns"]
        table_info.append(f"- {alias}表 ({display_name}): 列=[{', '.join(columns[:10])}{'...' if len(columns) > 10 else ''}]")
    
    tables_desc = '\n'.join(table_info)
    
    # 让 LLM 分析并拆解任务
    system_prompt = """你是一个数据分析任务规划器。你的任务是将用户的复杂需求拆解成多个可执行的步骤。

【输出格式要求】
严格按以下 JSON 格式输出，不要有任何其他文字：

```json
{
  "steps": [
    {
      "id": 1,
      "description": "对A表按xxx字段汇总",
      "input": ["A"],
      "output": "step_1"
    },
    {
      "id": 2,
      "description": "从B表提取xxx映射关系",
      "input": ["B"],
      "output": "step_2"
    },
    {
      "id": 3,
      "description": "关联step_1和step_2的结果，按xxx汇总",
      "input": ["step_1", "step_2"],
      "output": "final"
    }
  ]
}
```

【规则】
1. 每个步骤应该是原子操作（筛选、汇总、合并等）
2. 步骤之间通过 input/output 建立依赖关系
3. input 可以是表别名（如 "A", "B"）或前序步骤的 output（如 "step_1"）
4. 最后一步的 output 应该是 "final"
5. description 要清晰描述这一步做什么"""

    prompt = f"""## 可用数据表
{tables_desc}

## 用户需求
{query}

请分析这个需求，将其拆解成多个执行步骤。"""

    response = call_llm(prompt, system_prompt)
    
    # 提取 JSON
    import re
    json_match = re.search(r'```json\s*(.*?)\s*```', response, re.DOTALL)
    if not json_match:
        json_match = re.search(r'\{[\s\S]*"steps"[\s\S]*\}', response)
    
    if json_match:
        try:
            json_str = json_match.group(1) if '```' in response else json_match.group(0)
            plan_data = json.loads(json_str)
            
            steps = []
            for step in plan_data.get("steps", []):
                steps.append({
                    "id": step.get("id", len(steps) + 1),
                    "description": step.get("description", ""),
                    "input": step.get("input", []),
                    "output": step.get("output", f"step_{len(steps) + 1}"),
                })
            
            return {
                "is_multi_step": True,
                "steps": steps,
                "original_query": query,
                "message": f"检测到复杂任务，已拆解为 {len(steps)} 个步骤",
            }
        except json.JSONDecodeError:
            pass
    
    # 解析失败，降级为简单任务
    return {
        "is_multi_step": False,
        "message": "任务分析完成，可以直接执行",
        "fallback": True,
    }


@app.post("/api/execute-step")
async def execute_step(request: ExecuteStepRequest):
    """
    执行单个步骤，并存储中间结果
    """
    if not loaded_tables:
        raise HTTPException(status_code=400, detail="请先上传表格文件")
    
    # 收集输入数据
    input_dfs = {}
    for source in request.input_sources:
        if source.startswith("step_") or source == "final":
            # 从中间结果获取
            if source in step_results:
                input_dfs[f"df_{source}"] = step_results[source]
            else:
                return {
                    "success": False,
                    "error": f"中间结果 '{source}' 不存在，请先执行前序步骤",
                }
        else:
            # 从已加载表获取（通过别名）
            table_name = alias_to_table.get(source)
            if table_name and table_name in loaded_tables:
                input_dfs[f"df_{source}"] = loaded_tables[table_name]["df"]
            else:
                return {
                    "success": False,
                    "error": f"表 '{source}' 不存在",
                }
    
    # 构建提示词让 LLM 生成代码
    input_desc = []
    for var_name, df in input_dfs.items():
        cols = df.columns.tolist()
        sample = safe_dataframe_to_dict(df.head(2))
        input_desc.append(f"变量 `{var_name}`: 列={cols[:8]}, 示例={sample}")
    
    system_prompt = """你是 pandas 代码生成器。
【规则】
1. 只输出 Python 代码，用 ```python 包裹
2. 使用提供的变量名（df_A, df_B, df_step_1 等）
3. 结果存入 `result` 变量
4. 不要 import，pd/np 已预加载
5. 不要 print 或文件操作"""

    prompt = f"""## 可用数据
{chr(10).join(input_desc)}

## 任务描述
{request.step_description}

生成代码："""

    response = call_llm(prompt, system_prompt)
    code = extract_code(response)
    
    # 执行代码
    success, result, error = execute_code(code, input_dfs)
    
    if not success:
        return {
            "success": False,
            "error": error,
            "code": code,
        }
    
    if result is None or not isinstance(result, pd.DataFrame):
        return {
            "success": False,
            "error": "步骤执行完成，但结果不是 DataFrame",
            "code": code,
        }
    
    # 存储中间结果
    step_results[request.output_id] = result
    
    return {
        "success": True,
        "step_id": request.step_id,
        "output_id": request.output_id,
        "code": code,
        "result": {
            "type": "dataframe",
            "columns": result.columns.tolist(),
            "data": safe_dataframe_to_dict(result.head(50)),
            "total_rows": len(result),
        },
        "message": f"步骤 {request.step_id} 执行成功，结果已存储为 {request.output_id}",
    }


@app.post("/api/clear-step-results")
async def clear_step_results():
    """清除所有步骤中间结果"""
    step_results.clear()
    return {"success": True, "message": "已清除所有中间结果"}


@app.get("/api/step-results")
async def get_step_results():
    """获取当前所有中间结果的概览"""
    results = []
    for output_id, df in step_results.items():
        results.append({
            "id": output_id,
            "columns": df.columns.tolist(),
            "row_count": len(df),
        })
    return {"results": results}


@app.post("/api/save-step-as-table")
async def save_step_as_table(output_id: str, nickname: str = ""):
    """将步骤结果保存为正式表格"""
    if output_id not in step_results:
        raise HTTPException(status_code=400, detail=f"中间结果 '{output_id}' 不存在")
    
    df = step_results[output_id]
    
    # 分配别名
    alias = get_next_alias()
    table_name = f"结果_{output_id}"
    
    # 存储
    loaded_tables[table_name] = {
        "df": df,
        "path": "",
        "meta": {
            "table_name": table_name,
            "columns": df.columns.tolist(),
            "dtypes": {col: str(df[col].dtype) for col in df.columns},
            "row_count": len(df),
            "sample": safe_dataframe_to_dict(df.head(3)),
            "sheets": [],
            "current_sheet": "",
            "header_row": 1,
        },
        "alias": alias,
        "nickname": nickname if nickname else f"步骤{output_id}结果",
    }
    alias_to_table[alias] = table_name
    tables_meta[table_name] = loaded_tables[table_name]["meta"]
    
    return {
        "success": True,
        "alias": alias,
        "table_name": table_name,
        "message": f"已将 {output_id} 保存为 {alias}表",
    }


# ===== 导出功能 =====
class ExportRequest(BaseModel):
    data: List[Dict[str, Any]]
    file_name: Optional[str] = "报表导出"
    include_chart: bool = True
    chart_type: str = "bar"  # bar, pie, line
    label_column: Optional[str] = None
    value_column: Optional[str] = None


@app.post("/api/export")
async def export_to_excel(request: ExportRequest):
    """
    导出数据到 Excel（可选包含图表）
    """
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    import tempfile
    import uuid
    
    try:
        if not request.data:
            raise HTTPException(status_code=400, detail="没有数据可导出")
        
        # 转换为 DataFrame
        df = pd.DataFrame(request.data)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "数据"
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 添加图表
        if request.include_chart and len(df) >= 2:
            columns = df.columns.tolist()
            
            # 自动检测标签列和数值列
            label_col = request.label_column
            value_col = request.value_column
            
            if not label_col:
                # 找第一个非数值列
                for col in columns:
                    if df[col].dtype == 'object' or not np.issubdtype(df[col].dtype, np.number):
                        label_col = col
                        break
                if not label_col:
                    label_col = columns[0]
            
            if not value_col:
                # 找第一个数值列
                for col in columns:
                    if col != label_col and np.issubdtype(df[col].dtype, np.number):
                        value_col = col
                        break
            
            if value_col:
                label_col_idx = columns.index(label_col) + 1 if label_col in columns else 1
                value_col_idx = columns.index(value_col) + 1
                
                # 创建图表
                if request.chart_type == "pie":
                    chart = PieChart()
                    chart.title = f"{value_col} 占比"
                elif request.chart_type == "line":
                    chart = LineChart()
                    chart.title = f"{value_col} 趋势"
                    chart.y_axis.title = value_col
                    chart.x_axis.title = label_col
                else:
                    chart = BarChart()
                    chart.title = f"{value_col} 分布"
                    chart.y_axis.title = value_col
                    chart.x_axis.title = label_col
                
                # 数据范围
                data_ref = Reference(ws, min_col=value_col_idx, min_row=1, max_row=len(df) + 1)
                cats_ref = Reference(ws, min_col=label_col_idx, min_row=2, max_row=len(df) + 1)
                
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                
                # 放置图表
                ws.add_chart(chart, f"{chr(65 + len(columns) + 1)}2")
        
        # 保存到临时文件
        temp_dir = Path("./temp/exports")
        temp_dir.mkdir(parents=True, exist_ok=True)
        
        file_id = str(uuid.uuid4())[:8]
        file_name = f"{request.file_name}_{file_id}.xlsx"
        file_path = temp_dir / file_name
        
        wb.save(str(file_path))
        
        return {
            "success": True,
            "file_path": str(file_path),
            "file_name": file_name,
            "download_url": f"/api/download/{file_name}",
        }
    
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@app.get("/api/download/{file_name}")
async def download_file(file_name: str):
    """
    下载导出的文件
    """
    from fastapi.responses import FileResponse
    
    file_path = Path(f"./temp/exports/{file_name}")
    
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="文件不存在")
    
    return FileResponse(
        path=str(file_path),
        filename=file_name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# ===== 多 Sheet 操作 =====
class MergeToSheetRequest(BaseModel):
    source_table: str           # 源表名
    target_table: str           # 目标表名
    new_sheet_name: str         # 新 Sheet 名称
    include_chart: bool = False # 是否包含图表
    chart_type: str = "bar"     # 图表类型


@app.post("/api/merge-to-sheet")
async def merge_to_sheet(request: MergeToSheetRequest):
    """
    将源表数据合并到目标表的新 Sheet 中
    """
    from openpyxl import load_workbook, Workbook
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils.dataframe import dataframe_to_rows
    import uuid
    import shutil
    
    try:
        # 验证表是否存在
        if request.source_table not in loaded_tables:
            raise HTTPException(status_code=400, detail=f"源表 '{request.source_table}' 不存在")
        
        if request.target_table not in loaded_tables:
            raise HTTPException(status_code=400, detail=f"目标表 '{request.target_table}' 不存在")
        
        source_df = loaded_tables[request.source_table]["df"]
        target_path = loaded_tables[request.target_table]["path"]
        
        # 检查目标文件是否是 Excel
        if not target_path.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="目标文件必须是 Excel 格式")
        
        # 创建导出副本
        temp_dir = Path("./temp/exports")
        temp_dir.mkdir(parents=True, exist_ok=True)
        
        file_id = str(uuid.uuid4())[:8]
        output_name = f"{Path(target_path).stem}_merged_{file_id}.xlsx"
        output_path = temp_dir / output_name
        
        # 复制目标文件
        shutil.copy(target_path, output_path)
        
        # 打开并添加新 Sheet
        wb = load_workbook(str(output_path))
        
        # 检查 sheet 名是否已存在
        sheet_name = request.new_sheet_name
        if sheet_name in wb.sheetnames:
            # 如果存在，添加后缀
            counter = 1
            while f"{sheet_name}_{counter}" in wb.sheetnames:
                counter += 1
            sheet_name = f"{sheet_name}_{counter}"
        
        # 创建新 Sheet
        ws = wb.create_sheet(title=sheet_name)
        
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(source_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # 处理特殊值
                if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 添加图表（如果需要）
        if request.include_chart and len(source_df) >= 2:
            columns = source_df.columns.tolist()
            
            # 自动检测标签列和数值列
            label_col = None
            value_col = None
            
            for col in columns:
                if label_col is None and source_df[col].dtype == 'object':
                    label_col = col
                if value_col is None and np.issubdtype(source_df[col].dtype, np.number):
                    value_col = col
            
            if not label_col:
                label_col = columns[0]
            
            if value_col:
                label_col_idx = columns.index(label_col) + 1
                value_col_idx = columns.index(value_col) + 1
                
                if request.chart_type == "pie":
                    chart = PieChart()
                    chart.title = f"{value_col} 占比"
                else:
                    chart = BarChart()
                    chart.title = f"{value_col} 分布"
                
                data_ref = Reference(ws, min_col=value_col_idx, min_row=1, max_row=len(source_df) + 1)
                cats_ref = Reference(ws, min_col=label_col_idx, min_row=2, max_row=len(source_df) + 1)
                
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                
                ws.add_chart(chart, f"{chr(65 + len(columns) + 1)}2")
        
        # 保存
        wb.save(str(output_path))
        
        return {
            "success": True,
            "message": f"已将 '{request.source_table}' 合并到新 Sheet '{sheet_name}'",
            "file_name": output_name,
            "sheet_name": sheet_name,
            "download_url": f"/api/download/{output_name}",
        }
    
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"合并失败: {str(e)}")


class ExportMultiSheetRequest(BaseModel):
    tables: List[str]           # 要导出的表名列表
    sheet_names: List[str]      # 对应的 Sheet 名称
    file_name: str = "多表导出"


@app.post("/api/export-multi-sheet")
async def export_multi_sheet(request: ExportMultiSheetRequest):
    """
    将多个表导出到一个 Excel 的多个 Sheet
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import uuid
    
    try:
        if len(request.tables) != len(request.sheet_names):
            raise HTTPException(status_code=400, detail="表数量与 Sheet 名称数量不匹配")
        
        if not request.tables:
            raise HTTPException(status_code=400, detail="请至少选择一个表")
        
        # 验证所有表都存在
        for table in request.tables:
            if table not in loaded_tables:
                raise HTTPException(status_code=400, detail=f"表 '{table}' 不存在")
        
        # 创建工作簿
        wb = Workbook()
        
        # 删除默认 sheet
        default_sheet = wb.active
        
        for i, (table_name, sheet_name) in enumerate(zip(request.tables, request.sheet_names)):
            df = loaded_tables[table_name]["df"]
            
            if i == 0:
                ws = default_sheet
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    if isinstance(value, float) and (np.isnan(value) or np.isinf(value)):
                        value = None
                    ws.cell(row=r_idx, column=c_idx, value=value)
        
        # 保存
        temp_dir = Path("./temp/exports")
        temp_dir.mkdir(parents=True, exist_ok=True)
        
        file_id = str(uuid.uuid4())[:8]
        output_name = f"{request.file_name}_{file_id}.xlsx"
        output_path = temp_dir / output_name
        
        wb.save(str(output_path))
        
        return {
            "success": True,
            "message": f"已导出 {len(request.tables)} 个表到 Excel",
            "file_name": output_name,
            "sheets": request.sheet_names,
            "download_url": f"/api/download/{output_name}",
        }
    
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


class SplitToSheetsRequest(BaseModel):
    table_name: str          # 要拆分的表名
    split_column: str        # 拆分依据的列名
    file_name: str = "拆分导出"


@app.post("/api/split-to-sheets")
async def split_to_sheets(request: SplitToSheetsRequest):
    """
    按指定列的值将表格拆分成多个 Sheet 导出
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    import uuid
    
    try:
        if request.table_name not in loaded_tables:
            raise HTTPException(status_code=404, detail=f"表 '{request.table_name}' 不存在")
        
        df = loaded_tables[request.table_name]["df"]
        
        if request.split_column not in df.columns:
            raise HTTPException(status_code=400, detail=f"列 '{request.split_column}' 不存在")
        
        # 获取唯一值
        unique_values = df[request.split_column].dropna().unique()
        
        if len(unique_values) == 0:
            raise HTTPException(status_code=400, detail=f"列 '{request.split_column}' 没有有效值")
        
        if len(unique_values) > 100:
            raise HTTPException(status_code=400, detail=f"拆分值过多 ({len(unique_values)} 个)，最多支持 100 个 Sheet")
        
        # 创建工作簿
        wb = Workbook()
        default_sheet = wb.active
        sheet_names = []
        
        for i, value in enumerate(unique_values):
            # 过滤数据
            sub_df = df[df[request.split_column] == value]
            
            # 清理 Sheet 名称（Excel 限制 31 字符，不能包含特殊字符）
            sheet_name = str(value)[:31]
            for char in ['\\', '/', '*', '?', ':', '[', ']']:
                sheet_name = sheet_name.replace(char, '_')
            
            # 确保 Sheet 名称唯一
            original_name = sheet_name
            counter = 1
            while sheet_name in sheet_names:
                sheet_name = f"{original_name[:28]}_{counter}"
                counter += 1
            sheet_names.append(sheet_name)
            
            if i == 0:
                ws = default_sheet
                ws.title = sheet_name
            else:
                ws = wb.create_sheet(title=sheet_name)
            
            # 写入数据
            for r_idx, row in enumerate(dataframe_to_rows(sub_df, index=False, header=True), 1):
                for c_idx, col_value in enumerate(row, 1):
                    if isinstance(col_value, float) and (np.isnan(col_value) or np.isinf(col_value)):
                        col_value = None
                    ws.cell(row=r_idx, column=c_idx, value=col_value)
        
        # 保存
        temp_dir = Path("./temp/exports")
        temp_dir.mkdir(parents=True, exist_ok=True)
        
        file_id = str(uuid.uuid4())[:8]
        output_name = f"{request.file_name}_{file_id}.xlsx"
        output_path = temp_dir / output_name
        
        wb.save(str(output_path))
        
        return {
            "success": True,
            "message": f"已按 '{request.split_column}' 拆分成 {len(sheet_names)} 个 Sheet",
            "file_name": output_name,
            "sheets": sheet_names,
            "sheet_count": len(sheet_names),
            "download_url": f"/api/download/{output_name}",
        }
    
    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"拆分导出失败: {str(e)}")


@app.get("/api/table-columns/{table_name}")
async def get_table_columns(table_name: str):
    """获取表格的列名列表（用于拆分选择）"""
    if table_name not in loaded_tables:
        raise HTTPException(status_code=404, detail=f"表 '{table_name}' 不存在")
    
    df = loaded_tables[table_name]["df"]
    columns_info = []
    
    for col in df.columns:
        unique_count = df[col].nunique()
        columns_info.append({
            "name": col,
            "unique_count": unique_count,
            "suitable_for_split": 1 < unique_count <= 100  # 适合拆分的列
        })
    
    return {
        "success": True,
        "table_name": table_name,
        "columns": columns_info
    }


# ===== 报表工作区 API =====

@app.get("/api/workspace/status")
async def get_workspace_status():
    """获取当前工作区状态"""
    ws = workspace_manager.get_current()
    if not ws:
        return {
            "has_workspace": False,
            "message": "暂无活动报表"
        }
    
    return {
        "has_workspace": True,
        **ws.get_status()
    }


@app.post("/api/workspace/create")
async def create_workspace(name: str = "未命名报表"):
    """创建新的报表工作区"""
    ws = workspace_manager.create_workspace(name)
    return {
        "success": True,
        "message": f"已创建报表「{name}」",
        **ws.get_status()
    }


@app.post("/api/workspace/switch-sheet")
async def switch_active_sheet(sheet_name: str):
    """切换活动 Sheet"""
    ws = workspace_manager.get_current()
    if not ws:
        raise HTTPException(status_code=400, detail="没有活动的报表工作区")
    
    try:
        ws.switch_sheet(sheet_name)
        return {
            "success": True,
            "message": f"已切换到 Sheet「{sheet_name}」",
            "active_sheet": ws.active_sheet
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/api/workspace/sheet/{sheet_name}")
async def get_sheet_data(sheet_name: str, limit: int = 100):
    """获取指定 Sheet 的数据"""
    ws = workspace_manager.get_current()
    if not ws:
        raise HTTPException(status_code=400, detail="没有活动的报表工作区")
    
    try:
        df = ws.get_sheet_df(sheet_name)
        return {
            "success": True,
            "sheet_name": sheet_name,
            "columns": df.columns.tolist(),
            "data": safe_dataframe_to_dict(df.head(limit)),
            "total_rows": len(df)
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.post("/api/workspace/export")
async def export_workspace(file_name: Optional[str] = None):
    """导出当前报表工作区"""
    ws = workspace_manager.get_current()
    if not ws:
        raise HTTPException(status_code=400, detail="没有活动的报表工作区")
    
    if not ws.sheets:
        raise HTTPException(status_code=400, detail="报表中没有任何 Sheet")
    
    download_url = ws.export(file_name)
    
    return {
        "success": True,
        "message": f"报表已导出",
        "download_url": download_url,
        "sheets": list(ws.sheets.keys())
    }


# ===== 历史执行记录 API =====

class HistoryRecord(BaseModel):
    """历史记录模型"""
    id: int
    query: str              # 用户查询
    code: str               # 执行的代码
    analysis: str           # 分析说明
    timestamp: str          # 执行时间
    table_structure: Dict[str, List[str]]  # 表结构（表名 -> 列名列表）


def save_to_history(query: str, code: str, analysis: str):
    """保存执行记录到历史"""
    global history_id_counter
    history_id_counter += 1
    
    # 获取当前表结构
    table_structure = {}
    for name, info in loaded_tables.items():
        alias = info.get("alias", name)
        table_structure[alias] = info["meta"]["columns"]
    
    record = {
        "id": history_id_counter,
        "query": query,
        "code": code,
        "analysis": analysis,
        "timestamp": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"),
        "table_structure": table_structure,
    }
    
    execution_history.append(record)
    return record


def check_table_structure_match() -> tuple:
    """
    检查当前表结构是否与历史记录匹配
    返回: (是否匹配, 匹配的历史记录数, 不匹配原因)
    """
    if not execution_history:
        return False, 0, "没有历史执行记录"
    
    if not loaded_tables:
        return False, 0, "没有已加载的表格"
    
    # 获取当前表结构
    current_structure = {}
    for name, info in loaded_tables.items():
        alias = info.get("alias", name)
        current_structure[alias] = set(info["meta"]["columns"])
    
    # 检查第一条历史记录的表结构
    first_history = execution_history[0]
    history_structure = first_history.get("table_structure", {})
    
    # 比较表名
    if set(current_structure.keys()) != set(history_structure.keys()):
        return False, 0, f"表数量或名称不匹配。当前表: {list(current_structure.keys())}, 历史记录表: {list(history_structure.keys())}"
    
    # 比较每个表的列
    for alias in current_structure.keys():
        current_cols = current_structure[alias]
        history_cols = set(history_structure.get(alias, []))
        
        if current_cols != history_cols:
            missing = history_cols - current_cols
            extra = current_cols - history_cols
            msg_parts = []
            if missing:
                msg_parts.append(f"缺少列: {list(missing)}")
            if extra:
                msg_parts.append(f"多余列: {list(extra)}")
            return False, 0, f"表 {alias} 的列结构不匹配。{', '.join(msg_parts)}"
    
    return True, len(execution_history), "表结构匹配"


@app.get("/api/history")
async def get_history():
    """获取历史执行记录"""
    return {
        "success": True,
        "history": execution_history,
        "count": len(execution_history),
    }


@app.delete("/api/history/{record_id}")
async def delete_history_record(record_id: int):
    """删除指定的历史记录"""
    global execution_history
    
    original_length = len(execution_history)
    execution_history = [r for r in execution_history if r["id"] != record_id]
    
    if len(execution_history) == original_length:
        raise HTTPException(status_code=404, detail=f"历史记录 {record_id} 不存在")
    
    return {
        "success": True,
        "message": f"已删除历史记录 {record_id}",
        "remaining_count": len(execution_history),
    }


@app.delete("/api/history")
async def clear_history():
    """清空所有历史记录"""
    global execution_history, history_id_counter
    execution_history = []
    history_id_counter = 0
    
    return {
        "success": True,
        "message": "已清空所有历史记录",
    }


class BatchExecuteRequest(BaseModel):
    """批量执行请求"""
    record_ids: Optional[List[int]] = None  # 指定要执行的记录ID，为空则执行全部


@app.post("/api/history/batch-execute")
async def batch_execute_history(request: BatchExecuteRequest = None):
    """
    批量执行历史记录
    先检查表结构是否匹配，然后依次执行所有历史命令
    """
    if not execution_history:
        return {
            "status": "error",
            "message": "没有历史执行记录",
        }
    
    if not loaded_tables:
        return {
            "status": "error",
            "message": "请先上传表格文件",
        }
    
    # 检查表结构匹配
    is_match, count, reason = check_table_structure_match()
    
    if not is_match:
        return {
            "status": "structure_mismatch",
            "message": f"表结构不匹配: {reason}",
            "suggestion": "请确保上传的表格与历史记录中的表格结构一致（相同的表名和列名）",
        }
    
    # 准备执行
    records_to_execute = execution_history
    if request and request.record_ids:
        records_to_execute = [r for r in execution_history if r["id"] in request.record_ids]
    
    if not records_to_execute:
        return {
            "status": "error",
            "message": "没有找到要执行的记录",
        }
    
    # 准备 DataFrame 变量
    dataframes = {}
    for name, info in loaded_tables.items():
        alias = info.get("alias", name)
        dataframes[f"df_{alias}"] = info["df"]
        dataframes[f"df_{name}"] = info["df"]
    
    # 依次执行每条记录
    execution_logs = []
    final_result = None
    final_code = None
    
    for i, record in enumerate(records_to_execute):
        step_num = i + 1
        query = record["query"]
        code = record["code"]
        
        # 记录执行日志
        log_entry = {
            "step": step_num,
            "query": query,
            "code": code,
            "status": "pending",
        }
        
        try:
            success, result, error = execute_code(code, dataframes)
            
            if success:
                log_entry["status"] = "success"
                log_entry["message"] = f"执行成功"
                if isinstance(result, pd.DataFrame):
                    log_entry["result_rows"] = len(result)
                final_result = result
                final_code = code
            else:
                log_entry["status"] = "error"
                log_entry["message"] = error
                # 继续执行下一条（不中断）
        except Exception as e:
            log_entry["status"] = "error"
            log_entry["message"] = str(e)
        
        execution_logs.append(log_entry)
    
    # 构建响应
    success_count = sum(1 for log in execution_logs if log["status"] == "success")
    error_count = len(execution_logs) - success_count
    
    response = {
        "status": "success" if error_count == 0 else "partial_success",
        "message": f"执行完成: {success_count}/{len(execution_logs)} 条成功",
        "execution_logs": execution_logs,
        "total_steps": len(execution_logs),
        "success_count": success_count,
        "error_count": error_count,
    }
    
    # 添加最终结果
    if final_result is not None and isinstance(final_result, pd.DataFrame):
        response["result"] = {
            "type": "dataframe",
            "columns": final_result.columns.tolist(),
            "data": safe_dataframe_to_dict(final_result.head(100)),
            "total_rows": len(final_result),
        }
        response["code"] = final_code
    
    return response


@app.get("/api/history/check-structure")
async def check_structure():
    """检查当前表结构是否与历史记录匹配"""
    is_match, count, reason = check_table_structure_match()
    
    # 获取当前表结构信息
    current_structure = {}
    for name, info in loaded_tables.items():
        alias = info.get("alias", name)
        current_structure[alias] = info["meta"]["columns"]
    
    # 获取历史记录的表结构
    history_structure = {}
    if execution_history:
        history_structure = execution_history[0].get("table_structure", {})
    
    return {
        "is_match": is_match,
        "history_count": count,
        "message": reason,
        "current_structure": current_structure,
        "history_structure": history_structure,
    }


class TitleRequest(BaseModel):
    content: str
    max_length: int = 8


@app.post("/api/generate-title")
async def generate_title(req: TitleRequest):
    """根据用户消息生成简短标题（8字以内）"""
    try:
        prompt = f"""请根据以下内容生成一个简短的中文标题，要求：
1. 不超过{req.max_length}个字
2. 简洁概括主题
3. 只输出标题本身，不要其他内容

内容：{req.content[:200]}"""

        response = requests.post(
            SILICONFLOW_API_URL,
            headers={
                "Authorization": f"Bearer {SILICONFLOW_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": MODEL_NAME,
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": 50,
                "temperature": 0.7
            },
            timeout=10
        )
        
        if response.status_code == 200:
            result = response.json()
            title = result["choices"][0]["message"]["content"].strip()
            # 清理标题（去除引号、换行等）
            title = title.replace('"', '').replace("'", '').replace('\n', '').strip()
            # 限制长度
            if len(title) > req.max_length:
                title = title[:req.max_length]
            return {"title": title}
        else:
            # API 失败，使用简单截取
            return {"title": req.content[:req.max_length]}
            
    except Exception as e:
        # 出错时使用简单截取
        return {"title": req.content[:req.max_length]}


# ===== 启动服务 =====
if __name__ == "__main__":
    import uvicorn
    
    # 优先使用 BACKEND_PORT（Electron 传递），其次使用 PORT，默认 5000
    port = int(os.environ.get("BACKEND_PORT", os.environ.get("PORT", 5000)))
    
    print(f"🚀 启动 AI 报表生成器后端...")
    print(f"📍 地址: http://127.0.0.1:{port}")
    print(f"📚 API 文档: http://127.0.0.1:{port}/docs")
    
    # 注意：打包后不能用 "app:app" 字符串形式，必须直接传递 app 对象
    uvicorn.run(
        app,
        host="127.0.0.1",
        port=port,
        reload=False,
        log_level="info",
    )
